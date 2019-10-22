#coding:utf-8
import time

import databaseConnect
import writeExcel
import openpyxl
import xlrd;
from xlutils.copy import copy


def findRelationToConention(animals,conventions):
    # print animals;
    results = []
    for con in conventions:
        if con[7] == None:
            continue
        for animal in animals:
            if animal[1] == None:
                continue
            if animal[1] in con[7]:
                results.append([animal[0],animal[1],con[0],con[1],con[7]])
    return results
def findDescription():
    conn = databaseConnect.getConn()
    cur = conn.cursor()
    count = cur.execute('select 湿地编码,湿地名称,地理坐标,大致位置,海拔,面积,植物区系,动物区系 from 湿地公约')
    results = cur.fetchall()
    cur.close()
    conn.close()
    # for result in results:
    #     print result[7]
    return results

def animalTest():
    now = time.strftime("%M:%S")
    writeData(findRelationToConention(readData(),findDescription()))
    end = time.strftime("%M:%S")
    print (now+","+end)

def animalTest2():
    now = time.strftime("%M:%S")
    writeData2(findRelationToConention2(readData(),findDescription()))
    end = time.strftime("%M:%S")
    print (now+","+end)


def readData():
    data = xlrd.open_workbook("20180806.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(2,nrows):  #
        rows  = table.row_values(i) #行的数据放在数组里
        id = rows[0]
        sciName = rows[1]
        result.append([id,sciName])
    return result


def writeData(results):
    wb = openpyxl.load_workbook('result.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('result.xlsx')


def writeData2(results):
    wb = openpyxl.load_workbook('result2.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('result2.xlsx')


def writeDataTest():

    wb = openpyxl.load_workbook('result.xlsx')
    sheet = wb.active
    row_max = 1000
    print row_max
    for r in range(1, row_max):
        sheet.cell(row=r, column=1).value = 1

    wb.save('result.xlsx')


def processAniamlList(animalList):
    tempList = []
    for animal in animalList:
        tempList.append(animal)
    for animal in tempList:
        flag = True
        for a in tempList:
            if animal[1] == a[1]:
                continue
            if animal[1] in a[1]:
                flag = False
                break
        if flag==False:
            animalList.remove(animal)
    return animalList
def processAnimalListTest():
    animalList =[]
    animal1 = [1,'ab a']
    animal2 = [2,'ab']
    animal3 = [3,'a']
    animal4 = [4,'bbb']
    animalList.append(animal1)
    animalList.append(animal2)
    animalList.append(animal3)
    animalList.append(animal4)
    print processAniamlList(animalList)

def findRelationToConention2(animals,conventions):
    # print animals;
    results = []
    for con in conventions:
        if con[7] == None:
            continue
        animalList = []
        for animal in animals:
            if animal[1] == None:
                continue
            if animal[1] in con[7]:
                animalList.append([animal[0],animal[1]])
        animalList = processAniamlList(animalList)
        for animal in animalList:
            results.append([animal[0],animal[1],con[0],con[1],con[7]])
    return results


if __name__=="__main__":
    # print readData()
    #writeData()
    #writeDataTest()
    #animalTest()
    # processAnimalListTest()
    #animalTest2()
    print (len(findDescription()))