#coding:utf-8
import openpyxl
import xlrd;
import databaseConnect


def readAnimalData():
    data = xlrd.open_workbook("修改后动植物物种拉丁文_20180803_1.xls") #打开excel
    table = data.sheet_by_index(1)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(1,nrows):  #
        rows  = table.row_values(i) #行的数据放在数组里
        id = rows[0]
        sciName = rows[1]
        zhong = rows[2]
        result.append([id,sciName,zhong])
    return result

def readAnimalData0907():
    data = xlrd.open_workbook("animal20190907.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(1,nrows):  #
        rows  = table.row_values(i) #行的数据放在数组里
        id = rows[0]
        sciName = rows[1]
        zhong = rows[2]
        result.append([id,sciName,zhong])
    return result
def readPlantData():
    data = xlrd.open_workbook("plant_data_0918.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(0,nrows):  #
        rows  = table.row_values(i) #行的数据放在数组里
        id = rows[0]
        sciName = rows[1]
        result.append([id,sciName])
    return result

def findDescription():
    conn = databaseConnect.getConn()
    cur = conn.cursor()
    count = cur.execute('select 湿地编码,湿地名称,地理坐标,大致位置,海拔,面积,植物区系,动物区系 from 湿地公约')
    results = cur.fetchall()
    cur.close()
    conn.close()
    # for result in results:
    #     print result[7]
    return results


def findAnimalRelation(animals,conventions):
    # print animals;
    results = []
    for con in conventions:
        if con[7] == None:
            continue
        animalList = []
        for animal in animals:
            if animal[1] == None:
                continue
            if animal[1] in con[7]:
                animalList.append([animal[0],animal[1]])
        for animal in animalList:
            results.append([animal[0],animal[1],con[0],con[1],con[7]])
    return results

def findPlantRelation(plants,conventions):
    # print animals;
    results = []
    for con in conventions:
        if con[6] == None:
            continue
        plantlList = []
        for plant in plants:
            if plant[1] == None:
                continue
            if plant[1] in con[6]:
                plantlList.append([plant[0],plant[1]])
        plantlList = processAniamlList(plantlList)
        for plant in plantlList:
            results.append([plant[0],plant[1],con[0],con[1],con[6]])
    return results

def processAniamlList(animalList):

    tempList = []
    for animal in animalList:
        tempList.append(animal)
    for animal in tempList:
        flag = True
        for a in tempList:
            if animal[1] == a[1]:
                continue
            if animal[1] in a[1]:
                flag = False
                break
        if flag==False:
            animalList.remove(animal)
    return animalList

def processList(animalList):

    tempList = []
    for animal in animalList:
        tempList.append(animal)
    for animal in tempList:
        flag = True
        for a in tempList:
            if (animal[1] == a[1])&(animal[0]==a[0])&(animal[2]==a[2]):
                continue
            if animal[1] in a[1]:
                if(animal[0]!=a[0]):
                    flag = False
                    break
            if (animal[1]== a[1])&(animal[0]==a[0])&(len(animal[2])<len(a[2])):
                flag =False
                break
        if flag==False:
            animalList.remove(animal)
    return animalList

def findAnimal_copyDescription():
    conn = databaseConnect.getWetlandConn()
    cur = conn.cursor()
    count = cur.execute('select ScientificName,ID from animal_copy1')
    results = cur.fetchall()
    cur.close()
    conn.close()
    # for result in results:
    #     print result[7]
    return results


def writeNewData(animals,results):
    wb = openpyxl.load_workbook('无重复id_animal定稿.xlsx')
    sheet = wb.active
    row_max = len(results)

    print row_max
    for result in results:
        print result
    r = 1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r + 1
    wb.save('plantResults.xlsx')

def updateAnimalExcelId(animals,results):
    a=[]
    b=[]
    c=[]
    print len(results)
    for result in results:
        b.append(result[0])
    for animal in animals:
        if animal[1] in b:
            print ("in")
        else:
            c.append(animal[1])
            print (animal[1])
    for animal in animals:
        for result in results:
            # print animal[1]
            # print result[0]
            if animal[1] == result[0]:
                a.append([animal[0],result[1],animal[2]])
            else:
                continue
    print len(animals)
    print len(a)
    if len(animals)==len(a):
        return a
    else:
        return None

def reg(s):
    return s[0:s.find("(")]

def writePlantsData(results):
    wb = openpyxl.load_workbook('plant_wetland0918.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('plant_wetland0918.xlsx')

def writeAnimalsData(results):
    wb = openpyxl.load_workbook('animal_wetland0918.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('animal_wetland0918.xlsx')

def oneToone(results):
    length = len(results)
    list1 = []
    list2 = []
    for n in results:
        list1.append(n[0])
        list2.append(n[1])
    print (len(list1))
    print (len(list2))
    print (len(set(list1)))
    print (len(set(list2)))
    if len(set(list1))==length & len(set(list2))==length:
        return True
    else:
        return False
def findDup(results):
    for n in results:
        results.remove(n)
        for m in results:
            if n[0] == m[0]:
                print (n[0])

def oneTooneTest():
    testlist = []
    testlist.append([1,2])
    testlist.append([2,5,])
    testlist.append([3,1])

    return oneToone(testlist)

def findArticleDescription():
    conn = databaseConnect.getWetlandConn()
    cur = conn.cursor()
    count = cur.execute('select * from article')
    results = cur.fetchall()
    cur.close()
    conn.close()
    # for result in results:
    #     print result[7]
    return results

def findPlantFromLiterature(literature,plants):
    results = []
    for liter in literature:
        if liter == None:
            continue
        plantList = []
        for plant in plants:
            for w in liter:
                if w == None:
                    continue
                if isNum(w):
                    continue
                if plant==None:
                    continue
                if plant[1] in w:
                    plantList.append([plant[0],plant[1],w])
        plantList = processList(plantList)
        for plant in plantList:
            results.append([plant[0], plant[1], liter[0], liter[1],plant[2]])
            print(plant)
    return results


def findAnimalFromLiterature(literature,animals):
    results = []
    for liter in literature:
        if liter == None:
            continue
        animalList = []
        for animal in animals:
            for w in liter:
                if w == None:
                    continue
                if isNum(w):
                    continue
                if animal==None:
                    continue
                if animal[1] in w:
                    animalList.append([animal[0],animal[1],w])
        animalList = processList(animalList)
        for animal in animalList:
            results.append([animal[0], animal[1], liter[0], liter[1],animal[2]])
            print(animal)
    return results

def writeArticlePlants(results):
    wb = openpyxl.load_workbook('article_plant0918.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('article_plant0918.xlsx')

def writeArticleAnimals(results):
    wb = openpyxl.load_workbook('articleAnimalsResults.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('articleAnimalsResults.xlsx')
def isNum(value):
    try:
        value + 1
    except TypeError:
        return False
    else:
        return True

def processListTest():
    animalList =[]
    animal1 = [1,'ab a',"123"]
    animal2 = [2,'abd',"1234"]
    animal3 = [3,'abc',"1234"]

    animal4 = [4,'bbb',"12345"]
    animal5 = [4,'bbb',"123"]
    animalList.append(animal1)
    animalList.append(animal2)
    animalList.append(animal3)
    animalList.append(animal4)
    animalList.append(animal5)
    print processList(animalList)

if __name__=='__main__':

    #print (len(readAnimalData()))
    #print (len(readPlantData()))
    #findDup(readPlantData())
    #writePlantsData(findPlantRelation(readPlantData(),findDescription()))
    #writeAnimalsData(findAnimalRelation(readAnimalData0907(),findDescription()))
    #print (oneToone(readPlantData()))
    #print (oneToone(readAnimalData0907()))
    #print (updateAnimalExcelId(readAnimalData(),findAnimal_copyDescription()))
    #print (reg("Accipitridae(Hawks,Eagles)"))
    #print (len(readAnimalData0907()))
    #print(findPlantFromLiterature(findArticleDescription(),readPlantData()))
    writeArticlePlants(findPlantFromLiterature(findArticleDescription(),readPlantData()))
    #writeArticleAnimals(findAnimalFromLiterature(findArticleDescription(),readAnimalData0907()))
    #processListTest()