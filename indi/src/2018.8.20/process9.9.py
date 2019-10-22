#coding:utf-8
import openpyxl
import xlrd


def readAnimalData():
    data = xlrd.open_workbook("articleAnimalsResults的副本.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(1,nrows):  #
        rows  = table.row_values(i) #行的数据放在数组里
        id = rows[0]
        sciName = rows[1]
        article = rows[2]
        title= rows[3]
        des = rows[4]
        result.append([id,sciName,article,title,des])
    return result

def readPlantData():
    data = xlrd.open_workbook("articlePlantsResults的副本.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(1,nrows):  #
        rows = table.row_values(i)  # 行的数据放在数组里
        id = rows[0]
        sciName = rows[1]
        article = rows[2]
        title = rows[3]
        des = rows[4]
        result.append([id, sciName, article, title, des])
    return result

def delete(results):
    datas = [];
    for result in results:
        s = result[1]+' '
        # s2 = result[1]+','
        # s3 = result[1]+'.'
        if s in result[4]:
            datas.append([result[0],result[1],result[2],result[3],result[4]])
            continue
        # elif s2 in result[4]:
        #     datas.append([result[0], result[1], result[2], result[3], result[4]])
        #     continue
        # elif s3 in result[4]:
        #     datas.append([result[0], result[1], result[2], result[3], result[4]])
        #     continue
        else:
            continue
    return datas

def writePlantsData(results):
    wb = openpyxl.load_workbook('article_plants20190909.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('article_plants20190909.xlsx')

def writeAnimalsData(results):
    wb = openpyxl.load_workbook('article_animal20190909.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('article_animal20190909.xlsx')
if __name__=='__main__':
    #writePlantsData(delete(readPlantData()))
    writeAnimalsData(delete(readAnimalData()))