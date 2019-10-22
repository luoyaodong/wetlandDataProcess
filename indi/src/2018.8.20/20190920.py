#coding:utf-8
import openpyxl
import xlrd


def readAnimalFuleiData():
    data = xlrd.open_workbook("animal_dengji.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(0,nrows):  #
        rows = table.row_values(i)  # 行的数据放在数组里
        sciName = rows[0]
        parentName = rows[1]
        result.append([sciName,parentName])
    return result

def readPlantFuleiData():
    data = xlrd.open_workbook("plantdengji.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(0,nrows):  #
        rows = table.row_values(i)  # 行的数据放在数组里
        sciName = rows[0]
        parentName = rows[1]
        result.append([sciName,parentName])
    return result

def readWriteDengjiToArticle():
    data = xlrd.open_workbook("writeDengjiToArticle.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(0,nrows):  #
        rows = table.row_values(i)  # 行的数据放在数组里
        sciName = rows[0]
        parentName = rows[1]
        articleid = rows[2]
        title = rows[3]
        des = rows[4]
        result.append([sciName,parentName,articleid,title,des])
    print len(result)
    return result

def readPlantToArticle():
    data = xlrd.open_workbook("20190921异名.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(0,nrows):  #
        rows = table.row_values(i)  # 行的数据放在数组里
        sciName = rows[0]
        parentName = rows[1]
        articleid = rows[2]
        title = rows[3]
        des = rows[4]
        result.append([sciName,parentName,articleid,title,des])
    print len(result)
    return result


def getParents(animalFuleiData, param):
    parents=[]
    print param
    while (param!="Animalia")&(param!="Eulamellibranchia"):
        for animal in animalFuleiData:
            if animal[0] == param:
                parents.append(animal[1])
                param = animal[1]
                break
    return parents

def getPlantParents(plantFuleiData,param):
    parents = []
    print param
    while(param!="Plantae")&(param!="Bryophyta")&(param!="Fungi")&(param!="Typhoides"):
        for plant in plantFuleiData:
            if plant[0] == param:
                if plant[1] == None:
                    print "None parent found"
                parents.append(plant[1])
                param = plant[1]
                break
    return parents


def deleteKe(animalFuleiData,animals):
    temp = animals
    for a in animals:
        list =[]
        for ani in animals:
            if (a[2]==ani[2])&(a[0]!=ani[0]):
                list.append(ani)
        parents = getParents(animalFuleiData,a[0])
        if list:
            for l in list:
                if parents:
                    for p in parents:
                        if l[0] == p:
                            temp.remove(l)

    return temp

def deletePlant(animalFuleiData,animals):
    temp = animals
    for a in animals:
        list =[]
        for ani in animals:
            if (a[2]==ani[2])&(a[0]!=ani[0]):
                list.append(ani)
        parents = getPlantParents(animalFuleiData,a[0])
        if list:
            for l in list:
                if parents:
                    for p in parents:
                        if l[0] == p:
                            temp.remove(l)

    return temp

def writeDengjiToArticle2(results):
    wb = openpyxl.load_workbook('writeDengjiToArticle去除父类.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    # for result in results:
    #     print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    print r
    wb.save('writeDengjiToArticle去除父类.xlsx')

def writePlantDengjiToArticle(results):
    wb = openpyxl.load_workbook('writePlantDengjiToArticle去除父类.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    # for result in results:
    #     print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    print r
    wb.save('writePlantDengjiToArticle去除父类.xlsx')
if __name__ =="__main__":
    #writeDengjiToArticle2(deleteKe(readAnimalFuleiData(),readWriteDengjiToArticle()))
    #deleteKe(readAnimalFuleiData(),readWriteDengjiToArticle())
    #print  getParents(readAnimalFuleiData(),"Anodonta")



    writePlantDengjiToArticle(deletePlant(readPlantFuleiData(),readPlantToArticle()))
    #print getPlantParents(readPlantFuleiData(),"Typhoides")
    #readAnimalFuleiData()
