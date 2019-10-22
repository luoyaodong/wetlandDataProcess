#coding:utf-8
import openpyxl
import xlrd

import databaseConnect


def readPlantYiming():
    data = xlrd.open_workbook("异名0918.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(0,nrows):  #
        rows  = table.row_values(i) #行的数据放在数组里
        id = rows[0]
        sciName = rows[1]
        yiming = rows[2]

        result.append([id,sciName,yiming])
    return result

def findDescription():
    conn = databaseConnect.getConn()
    cur = conn.cursor()
    count = cur.execute('select 湿地编码,湿地名称,地理坐标,大致位置,海拔,面积,植物区系,动物区系 from 湿地公约')
    results = cur.fetchall()
    cur.close()
    conn.close()
    # for result in results:
    #     print result[7]
    return results

def findPlantRelation(plants,conventions):
    # print animals;
    results = []
    for con in conventions:
        if con[6] == None:
            continue
        plantlList = []
        for plant in plants:
            if plant[2] == None:
                continue
            if plant[2] in con[6]:
                plantlList.append([plant[0],plant[1],plant[2]])
        plantlList = processAniamlList(plantlList)
        for plant in plantlList:
            results.append([plant[0],plant[1],plant[2],con[0],con[1],con[6]])
    return results

def processAniamlList(animalList):

    tempList = []
    for animal in animalList:
        tempList.append(animal)
    for animal in tempList:
        flag = True
        for a in tempList:
            if animal[1] == a[1]:
                continue
            if animal[1] in a[1]:
                flag = False
                break
        if flag==False:
            animalList.remove(animal)
    return animalList

#等级专用去重
def processAniamlList2(animalList):

    tempList = []
    for animal in animalList:
        tempList.append(animal)
    for animal in tempList:
        flag = True
        for a in tempList:
            if animal[0] == a[0]:
                continue
            if animal[0] in a[0]:
                flag = False
                break
        if flag==False:
            animalList.remove(animal)
    return animalList

def processList(animalList):

    tempList = []
    for animal in animalList:
        tempList.append(animal)
    for animal in tempList:
        flag = True
        for a in tempList:
            if (animal[1] == a[1])&(animal[0]==a[0])&(animal[2]==a[2]):
                continue
            if animal[1] in a[1]:
                if(animal[0]!=a[0]):
                    flag = False
                    break
            if (animal[1]== a[1])&(animal[0]==a[0])&(len(animal[2])<len(a[2])):
                flag =False
                break
        if flag==False:
            animalList.remove(animal)
    return animalList

def processList2(animalList):

    tempList = []
    for animal in animalList:
        tempList.append(animal)
    for animal in tempList:
        flag = True
        for a in tempList:
            if (animal[1] == a[1])&(animal[0]==a[0])&(animal[2]==a[2]):
                continue
            if animal[0] in a[0]:
                if(animal[1]!=a[1]):
                    flag = False
                    break
            if (animal[1]== a[1])&(animal[0]==a[0])&(len(animal[2])<len(a[2])):
                flag =False
                break
        if flag==False:
            animalList.remove(animal)
    return animalList


def writeArticlePlants(results):
    wb = openpyxl.load_workbook('yiming_article.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        sheet.cell(row=r, column=6).value = result[5]
        r = r+1
    wb.save('yiming_article.xlsx')
def writeDengjiToArticle(results):
    wb = openpyxl.load_workbook('writeDengjiToArticle.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('writeDengjiToArticle.xlsx')

def writePlantDengjiToArticle(results):
    wb = openpyxl.load_workbook('writePlantDengjiToArticle.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('writePlantDengjiToArticle.xlsx')

def writePlantsData(results):
    wb = openpyxl.load_workbook('yiming_wetland.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        sheet.cell(row=r, column=6).value = result[5]
        r = r+1
    wb.save('yiming_wetland.xlsx')


def writePlantsYimingData(results):
    wb = openpyxl.load_workbook('20190921异名.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('20190921异名.xlsx')

def writePlantsYimingToArticleData(results):
    wb = openpyxl.load_workbook('yiming_to_article.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        sheet.cell(row=r, column=6).value = result[5]
        r = r+1
    wb.save('yiming_to_article.xlsx')

def readPlantToWetlandMergeData():
    data = xlrd.open_workbook("yiming_wetlandHebing.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(0,nrows):  #
        rows = table.row_values(i)  # 行的数据放在数组里
        id = rows[0]
        sciName = rows[1]
        wetlandid = rows[2]
        name = rows[3]
        des = rows[4]
        result.append([id, sciName, wetlandid, name, des])
    return result
def readAnimalFuleiData():
    data = xlrd.open_workbook("animal_dengji.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(0,nrows):  #
        rows = table.row_values(i)  # 行的数据放在数组里
        sciName = rows[0]
        parentName = rows[1]
        result.append([sciName,parentName])
    return result
def readFuleiData():
    data = xlrd.open_workbook("plantdengji.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(0,nrows):  #
        rows = table.row_values(i)  # 行的数据放在数组里
        sciName = rows[0]
        parentName = rows[1]
        result.append([sciName,parentName])
    return result

def deleteParentData(merge,fulei):
    temp=[]

    for m in merge:
        temp.append(m)

    for f in fulei:
        list = []
        for m in merge:
            if f[0]==m[1]:
                list.append(m)
        if list:
            for l in list:
                mm =[]
                for m in merge:
                    if (m[2]==l[2])&(m[1]!=l[1]):
                        mm.append(m)
                if mm:
                    for parent in mm:

                        if m[1] == f[1]:
                            print ("")
                            temp.remove(parent)
                else:

                    continue
        else:
            continue
    return temp
def writePlantsDeleteParent(results):
    wb = openpyxl.load_workbook('plant_wetland_nonparent.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    # for result in results:
    #     print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('plant_wetland_nonparent.xlsx')


def readPlantToWetlandData():
    data = xlrd.open_workbook("plant_wetland0918.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(1,nrows):  #
        rows = table.row_values(i)  # 行的数据放在数组里
        id = rows[0]
        sciName = rows[1]
        wetlandid = rows[2]
        name = rows[3]
        des = rows[4]
        result.append([id, sciName, wetlandid, name, des])
    return result

def readYimingToWetlandData():
    data = xlrd.open_workbook("yiming_wetland.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(1,nrows):  #
        rows = table.row_values(i)  # 行的数据放在数组里
        id = rows[0]
        sciName = rows[1]
        yiming = rows[2]
        wetlandid = rows[3]
        name = rows[4]
        des = rows[5]
        result.append([id, sciName,yiming, wetlandid, name, des])
    return result

def mergeList(yiming,total):
    temp =[]
    for t in yiming:
        temp.append(t)
    for ym in yiming:
        for t in total:
            if (ym[1]==t[0])&(ym[3]==t[2]):
                temp.remove(ym);
    print len(temp)
    for ym in temp:
        total.append([ym[1],"",ym[3],ym[4],ym[5]])

    return total

def findPlantFromLiterature(literature,plants):
    results = []
    for liter in literature:
        if liter == None:
            continue
        plantList = []
        for plant in plants:
            for w in liter:
                if w == None:
                    continue
                if isNum(w):
                    continue
                if plant==None:
                    continue
                if plant[2] in w:
                    plantList.append([plant[0],plant[1],w,plant[2]])
        plantList = processList(plantList)
        for plant in plantList:
            results.append([plant[0], plant[1], plant[3],liter[0], liter[1],plant[2]])
            print(plant)
    return results

def findPlantDengjiFromLiterature(literature,plants):
    results = []
    for liter in literature:
        if liter == None:
            continue
        plantList = []
        for plant in plants:
            for w in liter:
                if w == None:
                    continue
                if isNum(w):
                    continue
                if plant==None:
                    continue
                if plant[0]+' ' in w:
                    plantList.append([plant[0],plant[1],w])
        plantList = processList2(plantList)
        for plant in plantList:
            results.append([plant[0], plant[1],liter[0], liter[1],plant[2]])
            # print(plant)
    return results

def isNum(value):
    try:
        value + 1
    except TypeError:
        return False
    else:
        return True

def findArticleDescription():
    conn = databaseConnect.getWetlandConn()
    cur = conn.cursor()
    count = cur.execute('select * from article')
    results = cur.fetchall()
    cur.close()
    conn.close()
    # for result in results:
    #     print result[7]
    return results

def writeDengji(results):
    wb = openpyxl.load_workbook('dengji_wetland.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('dengji_wetland.xlsx')

def writeAnimalDengji(results):
    wb = openpyxl.load_workbook('animal_dengji_result.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    for result in results:
        print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('animal_dengji_result.xlsx')

def findDengjiRelation(plants,conventions):
    # print animals;
    results = []
    for con in conventions:
        if con[6] == None:
            continue
        plantlList = []
        for plant in plants:
            if plant[0] == None:
                continue
            if plant[0] in con[6]:
                plantlList.append([plant[0],plant[1]])
        plantlList = processAniamlList2(plantlList)
        for plant in plantlList:
            results.append([plant[0],plant[1],con[0],con[1],con[6]])
    return results

def findAnimalDengjiRelation(animals,conventions):
    # print animals;
    results = []
    for con in conventions:
        if con[7] == None:
            continue
        animalList = []
        for animal in animals:
            if animal[0] == None:
                continue
            if animal[0] in con[7]:
                animalList.append([animal[0],animal[1]])
        animalList = processAniamlList2(animalList)
        for animal in animalList:
            results.append([animal[0],animal[1],con[0],con[1],con[7]])
    return results


def readDengjiData():
    data = xlrd.open_workbook("dengji_wetland.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(0,nrows):  #
        rows = table.row_values(i)  # 行的数据放在数组里
        sciName = rows[0]
        parentName = rows[1]
        wetlandid = rows[2]
        name = rows[3]
        des = rows[4]
        result.append([parentName, sciName, wetlandid, name, des])
    return result

def writeDengjiDeleteParent(results):
    wb = openpyxl.load_workbook('dengji_wetland_nonparent.xlsx')
    sheet = wb.active
    row_max = len(results)
    print row_max
    # for result in results:
    #     print result
    r=1;
    for result in results:
        sheet.cell(row=r, column=1).value = result[0]
        sheet.cell(row=r, column=2).value = result[1]
        sheet.cell(row=r, column=3).value = result[2]
        sheet.cell(row=r, column=4).value = result[3]
        sheet.cell(row=r, column=5).value = result[4]
        r = r+1
    wb.save('dengji_wetland_nonparent.xlsx')


def readPlantToArticle():
    data = xlrd.open_workbook("writePlantDengjiToArticle.xlsx") #打开excel
    table = data.sheet_by_index(0)
    # table = data.sheet_by_name("Sheet2")#读sheet
    nrows = table.nrows #获得行数
    result = []
    for i in range(0,nrows):  #
        rows = table.row_values(i)  # 行的数据放在数组里
        sciName = rows[0]
        parentName = rows[1]
        articleid = rows[2]
        title = rows[3]
        des = rows[4]
        result.append([sciName,parentName,articleid,title,des])
    print len(result)
    return result

if __name__=='__main__':
    writePlantsYimingData(mergeList(readYimingToWetlandData(),readPlantToArticle()))
    #writePlantsYimingToArticleData(findPlantFromLiterature(findArticleDescription(),readPlantYiming()))

    #writePlantsDeleteParent(deleteParentData(readDengjiData(),readFuleiData()))
    #writeDengjiDeleteParent(deleteParentData(readDengjiData(),readFuleiData()))



    #writePlantsData(findPlantRelation(readPlantYiming(), findDescription()))
    #writeArticlePlants(findPlantFromLiterature(findArticleDescription(), readFuleiData()))


    #20号最新的可以用的两个函数
    # writeDengjiToArticle(findPlantDengjiFromLiterature(findArticleDescription(),readAnimalFuleiData()))
    #writePlantDengjiToArticle(findPlantDengjiFromLiterature(findArticleDescription(),readFuleiData()))

    #匹配湿地的两个函数
    #writeDengji(findDengjiRelation(readFuleiData(),findDescription()))
    #writeAnimalDengji(findAnimalDengjiRelation(readAnimalFuleiData(),findDescription()))